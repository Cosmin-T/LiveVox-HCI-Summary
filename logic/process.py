#!/usr/bin/env python3
# process.py

from logic.util import *
from logic.logs import *
import pandas as pd
import os
import glob
from datetime import datetime, timedelta
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
import numpy as np
import re

conf_log()

def get_files(directory, file_name):
    matching_files = glob.glob(f'{directory}/**/*{file_name}*', recursive=True)
    if matching_files:
        for file_path in matching_files:
            logging.info(file_path)
            return matching_files
    else:
        logging.error('No matching files found')
        return None

def clear_sheet(file_name):
    try:
        empty_df = pd.DataFrame()
        with pd.ExcelWriter(file_name, engine='openpyxl') as writer:
            empty_df.to_excel(writer, index=False, sheet_name='Sheet1')

        logging.info(f'Cleared sheet 1 in file: {file_name}')
    except Exception as e:
        logging.error(f'Error clearing sheet 1 in file {file_name}: {e}')

def generate_service_name(file):
    match = re.search(r'(\w+)_([A-Za-z0-9]+)_([A-Za-z0-9]+)', os.path.basename(file))
    if match:
        return '_'.join(match.groups())
    else:
        return os.path.basename(file)

def append_files(directory, file_name):
    today = datetime.now()
    yesterday = today - timedelta(days=1)
    yesterday_str = yesterday.strftime('%m-%d-%Y')

    files = get_files(directory, file_name)
    if files is None:
        logging.error('No matching files found')
        return

    dfs = []
    for file in files:
        try:
            file_df = pd.read_excel(file, header=None, names=HEADERS)
            file_df['Service'] = generate_service_name(file)
            dfs.append(file_df)
            logging.info(f'Loaded files: {file}')
        except Exception as e:
            logging.error(f'Error reading file {file}: {e}')

    if dfs:
        df = pd.concat(dfs, ignore_index=True)
        total_row = df[df.apply(lambda row: 'Total' in row.values, axis=1)]
        print(total_row)

        if not total_row.empty:
            updated_file_name = f'Overall Average Ready Time - {yesterday_str}.xlsx'
            if updated_file_name:
                clear_sheet(updated_file_name)
            total_row.to_excel(updated_file_name, index=False, engine='openpyxl')
            logging.info(f'Saved Total row as {updated_file_name}')
        else:
            logging.info('Total row not found in DataFrame')
    else:
        logging.info('No files found to concatenate')

def append_new_data():
    today = datetime.now()
    yesterday = today - timedelta(days=1)
    yesterday_str = yesterday.strftime('%m-%d-%Y')

    average_ready_time_file = None
    for filename in os.listdir(FINAL_FILE_DIRECTORY):
        if filename.startswith(FINAL_FILE_PREIFX):
            average_ready_time_file = os.path.join(FINAL_FILE_DIRECTORY, filename)
            print("Found file:", average_ready_time_file)

    if average_ready_time_file is None:
        print("No file found with prefix:", FINAL_FILE_PREIFX)
        return

    try:
        df = pd.DataFrame(columns=RELEVANT_COLS)
        sheet1_data = pd.read_excel(average_ready_time_file, sheet_name='Sheet1')
        print("Found file:", average_ready_time_file)

        for col in RELEVANT_COLS:
            if col in sheet1_data.columns:
                if col != 'Service':
                    df[col] = pd.to_numeric(sheet1_data[col], errors='coerce')
                    print(f'{col} converted to numeric')

                    if '(Min)' in col:
                        new_col_name = col.replace('(Min)', '(Sec)')
                        df.rename(columns={col: new_col_name}, inplace=True)
                        df[new_col_name] *= 60
                        print(f'{col} renamed to {new_col_name}')
                else:
                    df[col] = sheet1_data[col]
                    print(f'{col} not found in sheet1_data.columns, but still processed.')
            else:
                print(f'{col} not in {sheet1_data.columns}')

        df['Date'] = yesterday_str
        df['Average Ready Time (Sec)'] = df['Ready (Sec)'] / df['Successful Op Transfer']
        if 'Average Ready Time (Min)' in df.columns:
            del df['Average Ready Time (Min)']

        for index, row in df.iterrows():
            if row['Successful Op Transfer'] != 0:
                df.at[index, 'Average Ready Time (Sec)'] = row['Average Ready Time (Sec)']
            else:
                df.at[index, 'Average Ready Time (Sec)'] = 0
        total_index = df.index.max() + 1
        df.loc[total_index, 'Service'] = 'TOTAL'
        for col in df.columns[2:]:
            df.loc[total_index, col] = df[col].iloc[:total_index].sum()

        non_zero_values = df.loc[:total_index - 1, 'Average Ready Time (Sec)']
        total_non_zero_count = non_zero_values.count()
        total_non_zero_sum = non_zero_values.sum()
        total_average_ready_time = total_non_zero_sum / total_non_zero_count
        df.loc[total_index, 'Average Ready Time (Sec)'] = total_average_ready_time
        df['Average Ready Time (Sec)'] = df['Average Ready Time (Sec)'].round(2)

        colors = ['#009c89'] * len(df['Service'])
        colors[total_index] = '#7d2300'
        chart_data = df[['Service', 'Average Ready Time (Sec)']]
        plt.figure(figsize=(10, 6))
        plt.barh(chart_data['Service'], chart_data['Average Ready Time (Sec)'], color=colors)
        for index, value in enumerate(chart_data['Average Ready Time (Sec)']):
            plt.text(value, index, str(value))
        plt.xlabel('Average Ready Time (Sec)')
        plt.ylabel('Service')
        plt.title('Average Ready Time by Service')
        plt.tight_layout()
        chart_path = 'average_ready_time_chart.png'
        plt.savefig(chart_path)
        plt.close()
        print("Created and saved the bar chart")

        with pd.ExcelWriter(average_ready_time_file, engine='openpyxl', mode='a') as writer:
            if 'ART' not in writer.book.sheetnames:
                df.to_excel(writer, sheet_name='ART', index=False)
                print("Appended data to the excel sheet ART")

            idx = writer.book.sheetnames.index('ART')
            ws = writer.book.worksheets[idx]

            img = Image(chart_path)
            cell = ws.cell(row=16, column=1)
            ws.add_image(img, cell.coordinate)
            writer.book.save(average_ready_time_file)
            print("Added chart image to the excel sheet")

            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.2
                ws.column_dimensions[column].width = adjusted_width

            writer.book.save(average_ready_time_file)
            print("Adjusted column widths in the excel sheet")

        logging.info('Average Ready Time Data appended successfully.')
    except Exception as e:
        logging.error('An error occurred: ' + str(e))

def delete_downloaded_files(directory, file_name):
    matching_files = glob.glob(f'{directory}/**/*{file_name}*', recursive=True)
    if matching_files:
        for file_path in matching_files:
            logging.info(file_path)
    else:
        logging.error('No matching files found')
        return None

    for f in matching_files:
        if os.path.exists(f):
            os.remove(f)
            logging.info(f'{f} removed.')

def move_older_files():
    today = datetime.now()
    yesterday = today - timedelta(days=1)
    yesterday_str = yesterday.strftime('%m-%d-%Y')
    print(f"Yesterday's date: {yesterday_str}")

    for filename in os.listdir(FINAL_FILE_DIRECTORY):
        if filename.startswith(FINAL_FILE_PREIFX):
            average_ready_time_file = os.path.join(FINAL_FILE_DIRECTORY, filename)
            print("Found file:", average_ready_time_file)
            if average_ready_time_file.endswith(f' - {yesterday_str}.xlsx'):
                logging.info(f'File skipped: {average_ready_time_file}')
                print(f'File skipped: {average_ready_time_file}')
            else:
                destination = os.path.join(FILE_HISTORY, filename)
                os.rename(average_ready_time_file, destination)
                logging.info(f'Moved file to: {destination}')
                print(f'Moved file to: {destination}')


